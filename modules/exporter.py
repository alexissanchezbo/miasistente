"""
Genera el archivo Excel de salida con 4 pestañas y formato profesional.
"""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Paleta de colores ────────────────────────────────────────────────────────
C_TITULO_BG   = "1A1A2E"   # Azul marino muy oscuro
C_TITULO_FG   = "FFFFFF"
C_HEADER_BG   = "16213E"   # Azul oscuro
C_HEADER_FG   = "FFFFFF"
C_SEC1_BG     = "0F3460"   # Azul medio (sección raíz 4, 5.1, 5.2)
C_SEC1_FG     = "FFFFFF"
C_SEC2_BG     = "1B4F72"   # Azul más claro (subsección)
C_SEC2_FG     = "FFFFFF"
C_SUBTOT_BG   = "154360"   # Azul enfatizado para subtotales
C_SUBTOT_FG   = "F0F3FF"
C_DETAIL_BG   = "FFFFFF"
C_DETAIL_FG   = "1A1A2E"
C_DETAIL_ALT  = "EBF5FB"   # Filas alternas
C_NEG_FG      = "C0392B"   # Rojo para negativos
C_OBS_ALERTA  = "FDEDEC"   # Fondo alerta
C_OBS_AVISO   = "FEFBD8"   # Fondo aviso
C_OBS_INFO    = "EAF4FB"   # Fondo info

FMT_MONEY   = '#,##0.00'
FMT_MONEY_N = '#,##0.00;[Red]-#,##0.00'
FMT_PCT     = '0.0%'
# ▲ rojo = sube · ▼ verde = baja · — = sin cambio
FMT_VAR_PCT = '[Red]"▲ "0.0%;[Green]"▼ "0.0%;"  —"'


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, name="Calibri", italic=italic)

def _align(horizontal="right", vertical="center", wrap=False):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)

def _thin_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(bottom=thin)


def _nivel_cuenta(cod):
    return str(cod).count(".")


def _write_pyg_sheet(ws, empresa, titulo, filas, value_cols, show_pct=True, pct_mode="income"):
    """
    Escribe una hoja del PyG con las filas y columnas dadas.
    filas      : lista de dicts {'tipo', 'cod', 'concepto', 'values', 'pct'}
    value_cols : lista de nombres de columnas de valor
    pct_mode   : "income"    → % sobre ingresos (default)
                 "variation" → variación % vs período inmediato anterior
                               (▲ rojo si sube · ▼ verde si baja)
    """
    # ── Detectar cuentas madre (tienen hijos en la lista) ───────────────────
    all_cods = {str(f.get("cod", "")) for f in filas if f.get("tipo") == "data"}
    parent_cods = {
        cod for cod in all_cods
        if any(c.startswith(cod + ".") for c in all_cods if c != cod)
    }

    # ── Encabezados ──────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 14

    total_cols = 2 + (len(value_cols) * (2 if show_pct else 1))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(1, 1, empresa)
    c.font = _font(bold=True, color=C_TITULO_FG, size=13)
    c.fill = _fill(C_TITULO_BG)
    c.alignment = _align("center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    c = ws.cell(2, 1, titulo)
    c.font = _font(bold=False, color=C_TITULO_FG, size=11)
    c.fill = _fill(C_HEADER_BG)
    c.alignment = _align("center")

    for col in range(1, total_cols + 1):
        ws.cell(3, col).fill = _fill(C_HEADER_BG)

    # ── Fila de nombres de columnas ─────────────────────────────────────────
    ws.row_dimensions[4].height = 28
    ws.cell(4, 1, "Cód.").font = _font(bold=True, color=C_HEADER_FG, size=9)
    ws.cell(4, 1).fill = _fill(C_HEADER_BG)
    ws.cell(4, 1).alignment = _align("center")

    ws.cell(4, 2, "Concepto").font = _font(bold=True, color=C_HEADER_FG, size=9)
    ws.cell(4, 2).fill = _fill(C_HEADER_BG)
    ws.cell(4, 2).alignment = _align("left")

    col_idx = 3
    for vc in value_cols:
        ws.merge_cells(start_row=4, start_column=col_idx, end_row=4,
                       end_column=col_idx + (1 if show_pct else 0))
        c = ws.cell(4, col_idx, str(vc))
        c.font = _font(bold=True, color=C_HEADER_FG, size=9)
        c.fill = _fill(C_HEADER_BG)
        c.alignment = _align("center")
        col_idx += 2 if show_pct else 1

    # ── Subencabezados $ / % (o Var.%) ──────────────────────────────────────
    if show_pct:
        ws.row_dimensions[5].height = 14
        ws.cell(5, 1).fill = _fill(C_HEADER_BG)
        ws.cell(5, 2).fill = _fill(C_HEADER_BG)
        col_idx = 3
        for vi, vc in enumerate(value_cols):
            c1 = ws.cell(5, col_idx, "$")
            c1.font = _font(bold=True, color=C_HEADER_FG, size=8, italic=True)
            c1.fill = _fill(C_HEADER_BG)
            c1.alignment = _align("center")

            # Primer período en modo variación: sin subencabezado de %
            if pct_mode == "variation" and vi == 0:
                lbl_pct = ""
            elif pct_mode == "variation":
                lbl_pct = "Var.%"
            else:
                lbl_pct = "%"

            c2 = ws.cell(5, col_idx + 1, lbl_pct)
            c2.font = _font(bold=True, color=C_HEADER_FG, size=8, italic=True)
            c2.fill = _fill(C_HEADER_BG)
            c2.alignment = _align("center")
            col_idx += 2
        data_start_row = 6
    else:
        data_start_row = 5

    # ── Datos ────────────────────────────────────────────────────────────────
    # Registrar columnas de variación para aplicar formato condicional al final
    var_pct_col_letters = []   # letras de columna Excel de cada Var.%

    alt = False
    for fila in filas:
        tipo     = fila.get("tipo", "data")
        cod      = str(fila.get("cod", ""))
        concepto = str(fila.get("concepto", ""))
        values   = fila.get("values", {})
        pcts     = fila.get("pct", {})
        nivel    = _nivel_cuenta(cod)

        row = ws.max_row + 1
        ws.row_dimensions[row].height = 14 if tipo == "data" else 16

        # ── Colores según tipo y nivel ───────────────────────────────────────
        if tipo == "subtotal":
            bg, fg, bold, size = C_SUBTOT_BG, C_SUBTOT_FG, True, 10
        elif nivel == 0:
            bg, fg, bold, size = C_SEC1_BG, C_SEC1_FG, True, 10
        elif nivel == 1:
            bg, fg, bold, size = C_SEC2_BG, C_SEC2_FG, True, 9
        elif nivel == 2:
            bg = "2E4057" if not alt else "3D566E"
            fg, bold, size = "DDEEFF", True, 9
        else:
            bg   = C_DETAIL_ALT if alt else C_DETAIL_BG
            fg   = C_DETAIL_FG
            # Cuenta madre en nivel de detalle → negrita
            bold = (cod in parent_cods)
            size = 9

        if tipo == "data" and nivel >= 3:
            alt = not alt

        # ── Código y concepto ────────────────────────────────────────────────
        c_cod = ws.cell(row, 1, cod if tipo == "data" else "")
        c_cod.font = _font(bold=bold, color=fg, size=size - 1)
        c_cod.fill = _fill(bg)
        c_cod.alignment = _align("left")

        indent = "  " * max(0, nivel - 2) if tipo == "data" else ""
        c_con = ws.cell(row, 2, indent + concepto)
        c_con.font = _font(bold=bold, color=fg, size=size)
        c_con.fill = _fill(bg)
        c_con.alignment = _align("left", wrap=True)

        # ── Valores ──────────────────────────────────────────────────────────
        col_idx = 3
        for vi, vc in enumerate(value_cols):
            val      = values.get(vc, 0.0) or 0.0
            pct      = pcts.get(vc, 0.0)   or 0.0

            val_fg = fg
            if val < 0 and nivel >= 2:
                val_fg = C_NEG_FG if bg in (C_DETAIL_BG, C_DETAIL_ALT) else "FFAAAA"

            c_val = ws.cell(row, col_idx, val)
            c_val.number_format = FMT_MONEY
            c_val.font  = _font(bold=bold, color=val_fg, size=size)
            c_val.fill  = _fill(bg)
            c_val.alignment = _align("right")

            if show_pct:
                pct_col_letter = get_column_letter(col_idx + 1)

                if pct_mode == "variation":
                    if vi == 0:
                        # Primer período: celda vacía
                        c_pct = ws.cell(row, col_idx + 1, None)
                        c_pct.number_format = "@"
                    else:
                        prev_vc  = value_cols[vi - 1]
                        prev_val = values.get(prev_vc, 0.0) or 0.0
                        if abs(prev_val) > 0.0001:
                            variation = (val - prev_val) / abs(prev_val)
                        else:
                            variation = None   # indeterminado

                        c_pct = ws.cell(row, col_idx + 1, variation)
                        c_pct.number_format = FMT_VAR_PCT if variation is not None else "@"

                        # Registrar la columna para el formato condicional de iconos
                        if pct_col_letter not in var_pct_col_letters:
                            var_pct_col_letters.append(pct_col_letter)

                    c_pct.font      = _font(bold=False, color=fg, size=size - 1)
                    c_pct.fill      = _fill(bg)
                    c_pct.alignment = _align("center")

                else:
                    # Modo normal: % sobre ingresos
                    c_pct = ws.cell(row, col_idx + 1, pct / 100)
                    c_pct.number_format = FMT_PCT
                    c_pct.font      = _font(bold=False, color=fg, size=size - 1, italic=True)
                    c_pct.fill      = _fill(bg)
                    c_pct.alignment = _align("right")

                col_idx += 2
            else:
                col_idx += 1

    # (formato condicional de íconos manejado por el número de formato FMT_VAR_PCT)

    # ── Anchos de columna ────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 46
    col_idx = 3
    for _ in value_cols:
        ws.column_dimensions[get_column_letter(col_idx)].width = 16
        if show_pct:
            ws.column_dimensions[get_column_letter(col_idx + 1)].width = 10
            col_idx += 2
        else:
            col_idx += 1

    ws.freeze_panes = ws.cell(data_start_row, 3)


def _write_balance_sheet(ws, empresa, titulo, df_bg):
    """
    Escribe el Estado de Situación Financiera (Balance General).
    df_bg: DataFrame con columnas Cod, Concepto, Total
    Estructura: cuentas 1=Activos, 2=Pasivos, 3=Patrimonio
    """
    # Colores por grupo raíz
    COLORES = {
        "1": ("0B3D91", "FFFFFF"),   # Activos → azul oscuro
        "2": ("6B1A1A", "FFFFFF"),   # Pasivos → burdeos
        "3": ("145A32", "FFFFFF"),   # Patrimonio → verde oscuro
    }
    C_SUB   = ("1C2833", "F0F3FF")  # Subtotales
    C_D_BG  = "FFFFFF"
    C_D_ALT = "F2F3F4"
    C_D_FG  = "1A1A2E"

    # ── Encabezados ──────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 14

    for r in (1, 2, 3):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws.cell(1, 1, empresa)
    c.font = _font(bold=True, color=C_TITULO_FG, size=13)
    c.fill = _fill(C_TITULO_BG); c.alignment = _align("center")

    c = ws.cell(2, 1, titulo)
    c.font = _font(bold=False, color=C_TITULO_FG, size=11)
    c.fill = _fill(C_HEADER_BG); c.alignment = _align("center")

    for col in range(1, 4):
        ws.cell(3, col).fill = _fill(C_HEADER_BG)

    ws.row_dimensions[4].height = 24
    for col, (lbl, w, al) in enumerate([
        ("Cód.", 12, "center"), ("Concepto", 52, "left"), ("Saldo", 18, "right")
    ], 1):
        c = ws.cell(4, col, lbl)
        c.font = _font(bold=True, color=C_HEADER_FG, size=10)
        c.fill = _fill(C_HEADER_BG)
        c.alignment = _align(al)
        ws.column_dimensions[get_column_letter(col)].width = w

    data_start = 5

    # ── Detectar cuentas madre ───────────────────────────────────────────────
    all_cods = set(df_bg["Cod"].astype(str).str.strip())
    parent_cods = {
        cod for cod in all_cods
        if any(c.startswith(cod + ".") for c in all_cods if c != cod)
    }

    # ── Acumuladores para subtotales ─────────────────────────────────────────
    totales = {"1": 0.0, "2": 0.0, "3": 0.0}

    alt = False
    grupo_actual = None

    for _, rec in df_bg.iterrows():
        cod      = str(rec["Cod"]).strip()
        concepto = str(rec["Concepto"]).strip()
        valor    = float(rec["Total"]) if pd.notna(rec["Total"]) else 0.0
        nivel    = cod.count(".")
        raiz     = cod[0] if cod else ""

        if raiz not in ("1", "2", "3"):
            continue

        # Separador de sección al cambiar de raíz
        if raiz != grupo_actual:
            if grupo_actual is not None:
                # Línea subtotal del grupo anterior
                lbl_sub = {"1": "TOTAL ACTIVOS", "2": "TOTAL PASIVOS", "3": "TOTAL PATRIMONIO"}.get(grupo_actual, "TOTAL")
                row_s = ws.max_row + 1
                ws.row_dimensions[row_s].height = 16
                bg_s, fg_s = C_SUB
                for col in range(1, 4):
                    ws.cell(row_s, col).fill = _fill(bg_s)
                ws.cell(row_s, 2, lbl_sub).font = _font(bold=True, color=fg_s, size=10)
                ws.cell(row_s, 2).fill = _fill(bg_s)
                ws.cell(row_s, 2).alignment = _align("left")
                c_tot = ws.cell(row_s, 3, totales[grupo_actual])
                c_tot.number_format = FMT_MONEY
                c_tot.font = _font(bold=True, color=fg_s, size=10)
                c_tot.fill = _fill(bg_s)
                c_tot.alignment = _align("right")

            grupo_actual = raiz

        # Acumular solo hojas
        if cod not in parent_cods:
            totales[raiz] = totales.get(raiz, 0.0) + valor

        # ── Estilo por nivel ─────────────────────────────────────────────────
        if nivel == 0:
            bg_r, fg_r = COLORES.get(raiz, ("333333", "FFFFFF"))
            bg, fg, bold, size = bg_r, fg_r, True, 11
        elif nivel == 1:
            bg_r, fg_r = COLORES.get(raiz, ("333333", "FFFFFF"))
            # Versión más clara del color raíz
            bg = {"1": "154360", "2": "922B21", "3": "1E8449"}.get(raiz, "333333")
            fg, bold, size = "FFFFFF", True, 10
        elif nivel == 2:
            bg = {"1": "1A5276", "2": "A93226", "3": "27AE60"}.get(raiz, "444444")
            fg, bold, size = "EAFAF1" if raiz == "3" else "DDEEFF" if raiz == "1" else "FDEDEC", True, 9
        else:
            bg   = C_D_ALT if alt else C_D_BG
            fg   = C_D_FG
            bold = cod in parent_cods
            size = 9
            alt  = not alt

        row = ws.max_row + 1
        ws.row_dimensions[row].height = 14 if nivel >= 2 else 16

        indent = "  " * max(0, nivel - 1) if nivel >= 2 else ""

        c_cod = ws.cell(row, 1, cod if nivel >= 2 else "")
        c_cod.font = _font(bold=bold, color=fg, size=size - 1)
        c_cod.fill = _fill(bg); c_cod.alignment = _align("left")

        c_con = ws.cell(row, 2, indent + concepto)
        c_con.font = _font(bold=bold, color=fg, size=size)
        c_con.fill = _fill(bg); c_con.alignment = _align("left", wrap=True)

        c_val = ws.cell(row, 3, valor)
        c_val.number_format = FMT_MONEY
        val_color = "FFAAAA" if valor < 0 else fg
        c_val.font = _font(bold=bold, color=val_color, size=size)
        c_val.fill = _fill(bg); c_val.alignment = _align("right")

    # Subtotal del último grupo
    if grupo_actual:
        lbl_sub = {"1": "TOTAL ACTIVOS", "2": "TOTAL PASIVOS", "3": "TOTAL PATRIMONIO"}.get(grupo_actual, "TOTAL")
        row_s = ws.max_row + 1
        ws.row_dimensions[row_s].height = 16
        bg_s, fg_s = C_SUB
        for col in range(1, 4):
            ws.cell(row_s, col).fill = _fill(bg_s)
        ws.cell(row_s, 2, lbl_sub).font = _font(bold=True, color=fg_s, size=10)
        ws.cell(row_s, 2).fill = _fill(bg_s)
        ws.cell(row_s, 2).alignment = _align("left")
        c_tot = ws.cell(row_s, 3, totales.get(grupo_actual, 0.0))
        c_tot.number_format = FMT_MONEY
        c_tot.font = _font(bold=True, color=fg_s, size=10)
        c_tot.fill = _fill(bg_s)
        c_tot.alignment = _align("right")

    # ── Línea final: TOTAL PASIVO + PATRIMONIO ───────────────────────────────
    total_pp = totales.get("2", 0.0) + totales.get("3", 0.0)
    row_f = ws.max_row + 1
    ws.row_dimensions[row_f].height = 18
    for col in range(1, 4):
        ws.cell(row_f, col).fill = _fill(C_TITULO_BG)
    ws.cell(row_f, 2, "TOTAL PASIVO + PATRIMONIO").font = _font(bold=True, color="FFFFFF", size=11)
    ws.cell(row_f, 2).fill = _fill(C_TITULO_BG)
    ws.cell(row_f, 2).alignment = _align("left")
    c_fp = ws.cell(row_f, 3, total_pp)
    c_fp.number_format = FMT_MONEY
    c_fp.font = _font(bold=True, color="FFD700", size=11)
    c_fp.fill = _fill(C_TITULO_BG)
    c_fp.alignment = _align("right")

    ws.freeze_panes = ws.cell(data_start, 1)


def _write_sin_asignacion_sheet(ws, df_mayor, empresa):
    """
    Escribe una pestaña con dos tablas:
    1. Transacciones sin Proyecto asignado
    2. Transacciones sin Centro de Costo asignado
    """
    C_SEC = "16213E"
    C_TBL = "1A3A5C"

    def _bloque(ws, titulo, df_filtrado, start_row):
        row = start_row

        # Título del bloque
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws.cell(row, 1, titulo)
        c.font = _font(bold=True, color="FFFFFF", size=11)
        c.fill = _fill(C_TBL)
        c.alignment = _align("left")
        ws.row_dimensions[row].height = 20
        row += 1

        if df_filtrado.empty:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.cell(row, 1, "✔ Sin transacciones pendientes de asignación.").font = _font(color="2ECC71", size=9)
            ws.cell(row, 1).alignment = _align("left")
            return row + 2

        # Encabezados de la tabla
        headers = ["Fecha", "Cuenta", "Tipo Doc.", "Nro. Doc.", "Tercero / Persona", "Valor ($)"]
        widths  = [14,       36,       12,          18,          36,                   16]
        ws.row_dimensions[row].height = 16
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row, col, h)
            c.font = _font(bold=True, color="FFFFFF", size=9)
            c.fill = _fill(C_SEC)
            c.alignment = _align("center" if col != 2 else "left")
            ws.column_dimensions[get_column_letter(col)].width = w
        row += 1

        alt = False
        total = 0.0
        for _, tr in df_filtrado.iterrows():
            bg  = "EBF5FB" if alt else "FFFFFF"
            alt = not alt
            ws.row_dimensions[row].height = 13

            fecha = tr.get("Fecha", "")
            if hasattr(fecha, "strftime"):
                fecha = fecha.strftime("%d/%m/%Y")

            vals = [
                str(fecha),
                f"{tr.get('Codigo','').strip()} · {tr.get('Cuenta','').strip()}"[:60],
                str(tr.get("TipoDoc", "") or ""),
                str(tr.get("NumDoc",  "") or ""),
                str(tr.get("Tercero", "") or ""),
                float(tr.get("Monto", 0) or 0),
            ]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row, col, v)
                c.fill = _fill(bg)
                c.font = _font(size=9)
                if col == 6:
                    c.number_format = FMT_MONEY
                    c.alignment = _align("right")
                    total += float(v)
                elif col == 1:
                    c.alignment = _align("center")
                else:
                    c.alignment = _align("left")
            row += 1

        # Fila de total
        ws.row_dimensions[row].height = 14
        for col in range(1, 7):
            ws.cell(row, col).fill = _fill("1C2833")
        ws.cell(row, 5, f"TOTAL  ({len(df_filtrado)} transacciones)").font = _font(bold=True, color="F0F3FF", size=9)
        ws.cell(row, 5).fill = _fill("1C2833"); ws.cell(row, 5).alignment = _align("right")
        c_tot = ws.cell(row, 6, total)
        c_tot.number_format = FMT_MONEY
        c_tot.font = _font(bold=True, color="FFD700", size=9)
        c_tot.fill = _fill("1C2833"); c_tot.alignment = _align("right")

        return row + 2

    # ── Encabezado de la hoja ────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16
    ws.merge_cells("A1:F1")
    c = ws.cell(1, 1, empresa)
    c.font = _font(bold=True, color=C_TITULO_FG, size=13)
    c.fill = _fill(C_TITULO_BG); c.alignment = _align("center")

    ws.merge_cells("A2:F2")
    c = ws.cell(2, 1, "Transacciones Sin Asignación  ·  Proyecto y Centro de Costo")
    c.font = _font(bold=False, color=C_TITULO_FG, size=11)
    c.fill = _fill(C_HEADER_BG); c.alignment = _align("center")

    ws.row_dimensions[3].height = 8
    for col in range(1, 7):
        ws.cell(3, col).fill = _fill(C_HEADER_BG)

    current_row = 4

    # ── Bloque 1: Sin Proyecto ───────────────────────────────────────────────
    sin_proy = df_mayor[
        df_mayor["Proyecto"].astype(str).str.strip().isin(["", "nan", "None"])
    ].copy() if not df_mayor.empty else pd.DataFrame()

    current_row = _bloque(ws, "⚠️  Sin Proyecto asignado", sin_proy, current_row)

    # ── Bloque 2: Sin Centro de Costo ────────────────────────────────────────
    sin_cc = df_mayor[
        df_mayor["CentroCosto"].astype(str).str.strip().isin(["", "nan", "None"])
    ].copy() if not df_mayor.empty else pd.DataFrame()

    current_row = _bloque(ws, "⚠️  Sin Centro de Costo asignado", sin_cc, current_row)

    ws.freeze_panes = "A4"


def _write_dashboard_sheet(
    ws, empresa, periodo_desc,
    filas_mes, value_cols_mes,
    filas_proyecto, value_cols_proyecto,
    filas_cc, value_cols_cc,
    df_balance=None,
):
    """Dashboard ejecutivo — primera pestaña del Excel."""

    # ── Helpers ──────────────────────────────────────────────────────────────
    def _sub(filas, label, col="TOTAL"):
        lab_up = label.upper()
        for f in filas:
            if f.get("tipo") == "subtotal" and lab_up in f.get("concepto", "").upper():
                return float(f.get("values", {}).get(col, 0) or 0)
        return 0.0

    def _sumcod(filas, prefix, col="TOTAL"):
        total = 0.0
        for f in filas:
            if f.get("tipo") == "data" and str(f.get("cod", "")).startswith(prefix):
                total += float(f.get("values", {}).get(col, 0) or 0)
        return total

    def _pct(val, base):
        return (val / abs(base)) if abs(base or 0) > 0.001 else 0.0

    # ── Colores locales ───────────────────────────────────────────────────────
    C_SEC_HDR = "1C2833"
    C_TBL_HDR = "16213E"
    C_TOT_ROW = "0F3460"
    C_ALT_LOC = "EBF5FB"

    NCOLS = 11

    # ── Anchos de columna ────────────────────────────────────────────────────
    for col_l, w in [("A",1),("B",42),("C",18),("D",18),("E",18),("F",10),
                     ("G",18),("H",18),("I",10),("J",18),("K",10)]:
        ws.column_dimensions[col_l].width = w

    # ══════════════════════════════════════════════════════════════════════════
    # TÍTULO
    # ══════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[1].height = 24
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
    c = ws.cell(1, 1, empresa)
    c.font = _font(bold=True, color=C_TITULO_FG, size=14)
    c.fill = _fill(C_TITULO_BG); c.alignment = _align("center")

    ws.row_dimensions[2].height = 18
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOLS)
    c = ws.cell(2, 1, f"Dashboard Ejecutivo   ·   {periodo_desc}")
    c.font = _font(bold=False, color=C_TITULO_FG, size=11)
    c.fill = _fill(C_HEADER_BG); c.alignment = _align("center")

    ws.row_dimensions[3].height = 6
    for col in range(1, NCOLS + 1):
        ws.cell(3, col).fill = _fill(C_HEADER_BG)

    row = 4

    # ══════════════════════════════════════════════════════════════════════════
    # SECCIÓN 1 — ESTADO DE RESULTADOS CONSOLIDADO
    # ══════════════════════════════════════════════════════════════════════════
    vcol = "TOTAL"
    ingr    = _sumcod(filas_proyecto, "4",       vcol)
    cv_mat  = _sumcod(filas_proyecto, "5.1.1",   vcol)
    cv_mod  = _sumcod(filas_proyecto, "5.1.2",   vcol)
    cv_moi  = _sumcod(filas_proyecto, "5.1.3",   vcol)
    cv_cif  = _sumcod(filas_proyecto, "5.1.4",   vcol)
    gv      = _sumcod(filas_proyecto, "5.2.1.1", vcol)
    ga      = _sumcod(filas_proyecto, "5.2.1.2", vcol)
    gf      = _sumcod(filas_proyecto, "5.2.1.3", vcol)
    g_nop   = (_sumcod(filas_proyecto, "5.2.2",  vcol) +
               _sumcod(filas_proyecto, "5.2.3",  vcol))
    impues  = _sumcod(filas_proyecto, "5.2.4",   vcol)
    ut_bruta = _sub(filas_proyecto, "UTILIDAD BRUTA",              vcol)
    ut_bi    = _sub(filas_proyecto, "UTILIDAD BRUTA INDUSTRIAL",   vcol)
    ut_op    = _sub(filas_proyecto, "UTILIDAD OPERACIONAL",        vcol)
    ut_ai    = _sub(filas_proyecto, "ANTES DE IMPUESTOS",          vcol)
    ut_neta  = _sub(filas_proyecto, "UTILIDAD NETA",               vcol)
    gastos_op = gv + ga

    # ── Encabezado sección ────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 20
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    c = ws.cell(row, 1, "    ESTADO DE RESULTADOS CONSOLIDADO")
    c.font = _font(bold=True, color="FFFFFF", size=11)
    c.fill = _fill(C_SEC_HDR); c.alignment = _align("left", vertical="center")
    row += 1

    # ── Cabecera de tabla (solo cubre cols A-D, resto fondo neutro) ──────────
    ws.row_dimensions[row].height = 16
    for col in range(1, NCOLS + 1):
        ws.cell(row, col).fill = _fill(C_TBL_HDR)
    for col_n, hdr, al in [
        (2, "Concepto",    "left"),
        (3, "Valor ($)",   "right"),
        (4, "% Ingresos",  "right"),
    ]:
        c = ws.cell(row, col_n, hdr)
        c.font = _font(bold=True, color="FFFFFF", size=10)
        c.fill = _fill(C_TBL_HDR); c.alignment = _align(al)
    row += 1

    # ── Función auxiliar para fila de dato ───────────────────────────────────
    def _drow(label, val, base, bold=False, bg="FFFFFF", dark_bg=False, indent=0):
        nonlocal row
        ws.row_dimensions[row].height = 15
        for col in range(1, NCOLS + 1):
            ws.cell(row, col).fill = _fill(bg)
        fg_txt = "FFFFFF" if dark_bg else "222222"
        fg_val = "FFFFFF" if dark_bg else ("C0392B" if val < 0 else "1A5276")
        fg_pct = "CCCCCC" if dark_bg else ("C0392B" if val < 0 else "888888")
        c = ws.cell(row, 2, ("    " * indent) + label)
        c.font = _font(bold=bold, color=fg_txt, size=10)
        c.fill = _fill(bg); c.alignment = _align("left")
        c_v = ws.cell(row, 3, val)
        c_v.number_format = FMT_MONEY
        c_v.font = _font(bold=bold, color=fg_val, size=10)
        c_v.fill = _fill(bg); c_v.alignment = _align("right")
        c_p = ws.cell(row, 4, _pct(val, base))
        c_p.number_format = FMT_PCT
        c_p.font = _font(bold=False, color=fg_pct, size=9, italic=True)
        c_p.fill = _fill(bg); c_p.alignment = _align("right")
        row += 1

    def _srow(label, val, base, bg=None, fg="FFFFFF"):
        nonlocal row
        bg = bg or C_SUBTOT_BG
        ws.row_dimensions[row].height = 18
        for col in range(1, NCOLS + 1):
            ws.cell(row, col).fill = _fill(bg)
        c = ws.cell(row, 2, label)
        c.font = _font(bold=True, color=fg, size=11)
        c.fill = _fill(bg); c.alignment = _align("left", vertical="center")
        c_v = ws.cell(row, 3, val)
        c_v.number_format = FMT_MONEY
        c_v.font = _font(bold=True, color="FFD700" if val >= 0 else "FF8080", size=11)
        c_v.fill = _fill(bg); c_v.alignment = _align("right", vertical="center")
        c_p = ws.cell(row, 4, _pct(val, base))
        c_p.number_format = FMT_PCT
        c_p.font = _font(bold=True, color="DDDDDD", size=10)
        c_p.fill = _fill(bg); c_p.alignment = _align("right", vertical="center")
        row += 1

    alt = False
    def _alt():
        nonlocal alt
        bg = C_ALT_LOC if alt else "FFFFFF"
        alt = not alt
        return bg

    _srow("INGRESOS NETOS",                  ingr,     ingr, bg="0B3D91")
    _drow("(-) Costo de Materias Primas",   -cv_mat,  ingr, indent=1, bg=_alt())
    _srow("UTILIDAD BRUTA",                 ut_bruta, ingr)
    _drow("(-) Mano de Obra Directa",       -cv_mod,  ingr, indent=1, bg=_alt())
    _drow("(-) Mano de Obra Indirecta",     -cv_moi,  ingr, indent=1, bg=_alt())
    _drow("(-) Costos Ind. Fabricacion",    -cv_cif,  ingr, indent=1, bg=_alt())
    _srow("UTILIDAD BRUTA INDUSTRIAL",      ut_bi,    ingr)
    _drow("(-) Gastos de Ventas",           -gv,      ingr, indent=1, bg=_alt())
    _drow("(-) Gastos Administrativos",     -ga,      ingr, indent=1, bg=_alt())
    _srow("UTILIDAD OPERACIONAL",           ut_op,    ingr)
    _drow("(-) Gastos Financieros",         -gf,      ingr, indent=1, bg=_alt())
    _srow("UTILIDAD ANTES DE IMPUESTOS",    ut_ai,    ingr)
    _drow("(-) Impuestos",                  -impues,  ingr, indent=1, bg=_alt())
    _srow("UTILIDAD NETA",                  ut_neta,  ingr, bg=C_TITULO_BG, fg="FFFFFF")

    row += 1  # spacer

    # ══════════════════════════════════════════════════════════════════════════
    # SECCIÓN 2 — SITUACIÓN FINANCIERA RESUMEN
    # ══════════════════════════════════════════════════════════════════════════
    if df_balance is not None and not df_balance.empty:
        tot_bg = {"1": 0.0, "2": 0.0, "3": 0.0}
        all_cods_b = set(df_balance["Cod"].astype(str).str.strip())
        parent_b = {
            cod for cod in all_cods_b
            if any(c2.startswith(cod + ".") for c2 in all_cods_b if c2 != cod)
        }
        for _, rec in df_balance.iterrows():
            cod = str(rec["Cod"]).strip()
            val = float(rec.get("Total", 0) or 0)
            raiz = cod[0] if cod else ""
            if raiz in tot_bg and cod not in parent_b:
                tot_bg[raiz] += val

        ws.row_dimensions[row].height = 16
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
        c = ws.cell(row, 1, "  SITUACIÓN FINANCIERA — RESUMEN")
        c.font = _font(bold=True, color="FFFFFF", size=10)
        c.fill = _fill(C_SEC_HDR); c.alignment = _align("left")
        row += 1

        for lbl, raiz, bg_b in [
            ("Total Activos",    "1", "0B3D91"),
            ("Total Pasivos",    "2", "6B1A1A"),
            ("Total Patrimonio", "3", "145A32"),
        ]:
            val_b = tot_bg[raiz]
            ws.row_dimensions[row].height = 14
            for col in range(1, NCOLS + 1):
                ws.cell(row, col).fill = _fill(bg_b)
            c = ws.cell(row, 2, lbl)
            c.font = _font(bold=True, color="FFFFFF", size=10)
            c.fill = _fill(bg_b); c.alignment = _align("left")
            c_v = ws.cell(row, 3, val_b)
            c_v.number_format = FMT_MONEY
            c_v.font = _font(bold=True, color="FFD700", size=10)
            c_v.fill = _fill(bg_b); c_v.alignment = _align("right")
            pct_b = _pct(val_b, tot_bg["1"])
            c_p = ws.cell(row, 4, pct_b)
            c_p.number_format = FMT_PCT
            c_p.font = _font(bold=False, color="CCCCCC", size=9, italic=True)
            c_p.fill = _fill(bg_b); c_p.alignment = _align("right")
            row += 1

        row += 1  # spacer

    # ══════════════════════════════════════════════════════════════════════════
    # SECCIÓN 3 — EVOLUCIÓN MENSUAL
    # ══════════════════════════════════════════════════════════════════════════
    if filas_mes and value_cols_mes:
        ws.row_dimensions[row].height = 16
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
        c = ws.cell(row, 1, "  EVOLUCIÓN MENSUAL")
        c.font = _font(bold=True, color="FFFFFF", size=10)
        c.fill = _fill(C_SEC_HDR); c.alignment = _align("left")
        row += 1

        # Cabecera tabla mensual (usa cols B a I)
        ws.row_dimensions[row].height = 14
        for col in range(1, NCOLS + 1):
            ws.cell(row, col).fill = _fill(C_TBL_HDR)
        for col_n, hdr, al in [
            (2,"Período","left"),(3,"Ingresos","right"),(4,"Ut. Bruta","right"),
            (5,"% MB","right"),(6,"Ut. Operacional","right"),(7,"% Op.","right"),
            (8,"Ut. Neta","right"),(9,"% Neta","right"),
        ]:
            c = ws.cell(row, col_n, hdr)
            c.font = _font(bold=True, color="FFFFFF", size=9)
            c.fill = _fill(C_TBL_HDR); c.alignment = _align(al)
        row += 1

        alt_m = False
        for vc in value_cols_mes:
            m_ingr = _sumcod(filas_mes, "4", vc)
            m_ub   = _sub(filas_mes, "UTILIDAD BRUTA", vc)
            m_uop  = _sub(filas_mes, "UTILIDAD OPERACIONAL", vc)
            m_un   = _sub(filas_mes, "UTILIDAD NETA", vc)
            is_tot = str(vc).upper() in ("TOTAL", "TOTALES", "TOT")
            bg_m   = C_TOT_ROW if is_tot else (C_ALT_LOC if alt_m else "FFFFFF")
            alt_m  = not alt_m

            ws.row_dimensions[row].height = 13
            for col in range(1, NCOLS + 1):
                ws.cell(row, col).fill = _fill(bg_m)

            fg_m  = "FFFFFF" if is_tot else "1A1A2E"
            c = ws.cell(row, 2, str(vc))
            c.font = _font(bold=is_tot, color=fg_m, size=9)
            c.fill = _fill(bg_m); c.alignment = _align("left")

            for col_n, val in [(3,m_ingr),(4,m_ub),(6,m_uop),(8,m_un)]:
                c_v = ws.cell(row, col_n, val)
                c_v.number_format = FMT_MONEY
                vfg = ("FFD700" if val >= 0 else "FF9090") if is_tot else ("C0392B" if val < 0 else fg_m)
                c_v.font = _font(bold=is_tot, color=vfg, size=9)
                c_v.fill = _fill(bg_m); c_v.alignment = _align("right")

            for col_n, val, base in [(5,m_ub,m_ingr),(7,m_uop,m_ingr),(9,m_un,m_ingr)]:
                pct_v = _pct(val, base)
                c_p = ws.cell(row, col_n, pct_v)
                c_p.number_format = FMT_PCT
                pfg = ("DDDDDD" if pct_v >= 0 else "FF9090") if is_tot else ("C0392B" if pct_v < 0 else "777777")
                c_p.font = _font(bold=is_tot, color=pfg, size=9, italic=not is_tot)
                c_p.fill = _fill(bg_m); c_p.alignment = _align("right")

            row += 1
        row += 1  # spacer

    # ══════════════════════════════════════════════════════════════════════════
    # SECCIÓN 4 — RESUMEN POR PROYECTO
    # ══════════════════════════════════════════════════════════════════════════
    proyectos = [c2 for c2 in value_cols_proyecto
                 if c2 not in ("TOTAL", "Sin Proyecto")]
    if proyectos:
        ws.row_dimensions[row].height = 16
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
        c = ws.cell(row, 1, f"  RESUMEN POR PROYECTO  ({len(proyectos)} proyectos detectados)")
        c.font = _font(bold=True, color="FFFFFF", size=10)
        c.fill = _fill(C_SEC_HDR); c.alignment = _align("left")
        row += 1

        ws.row_dimensions[row].height = 14
        for col in range(1, NCOLS + 1):
            ws.cell(row, col).fill = _fill(C_TBL_HDR)
        for col_n, hdr, al in [
            (2,"Proyecto","left"),(3,"Ingresos Netos","right"),(4,"Costo Ventas","right"),
            (5,"Ut. Bruta","right"),(6,"% MB","right"),(7,"Gastos Op.","right"),
            (8,"Ut. Operacional","right"),(9,"% Op.","right"),(10,"Ut. Neta","right"),(11,"% Neta","right"),
        ]:
            c = ws.cell(row, col_n, hdr)
            c.font = _font(bold=True, color="FFFFFF", size=9)
            c.fill = _fill(C_TBL_HDR); c.alignment = _align(al)
        ws.cell(row, 1).fill = _fill(C_TBL_HDR)
        row += 1

        # Recopilar y ordenar por ingresos desc
        proj_data = []
        for proy in proyectos:
            p_ingr = _sumcod(filas_proyecto, "4",       proy)
            p_cmat = _sumcod(filas_proyecto, "5.1.1",   proy)
            p_ub   = _sub(filas_proyecto,   "UTILIDAD BRUTA",        proy)
            p_gop  = (_sumcod(filas_proyecto, "5.2.1.1", proy) +
                      _sumcod(filas_proyecto, "5.2.1.2", proy))
            p_uop  = _sub(filas_proyecto,   "UTILIDAD OPERACIONAL",  proy)
            p_un   = _sub(filas_proyecto,   "UTILIDAD NETA",         proy)
            proj_data.append((proy, p_ingr, p_cmat, p_ub, p_gop, p_uop, p_un))
        proj_data.sort(key=lambda x: x[1], reverse=True)

        alt_p = False
        for proy, p_ingr, p_cmat, p_ub, p_gop, p_uop, p_un in proj_data:
            bg_p = C_ALT_LOC if alt_p else "FFFFFF"
            alt_p = not alt_p
            ws.row_dimensions[row].height = 13
            for col in range(1, NCOLS + 1):
                ws.cell(row, col).fill = _fill(bg_p)
            c = ws.cell(row, 2, proy)
            c.font = _font(bold=False, color="1A1A2E", size=9)
            c.fill = _fill(bg_p); c.alignment = _align("left")
            for col_n, val in [(3,p_ingr),(4,p_cmat),(5,p_ub),(7,p_gop),(8,p_uop),(10,p_un)]:
                c_v = ws.cell(row, col_n, val)
                c_v.number_format = FMT_MONEY
                c_v.font = _font(bold=False, color="C0392B" if val < 0 else "1A1A2E", size=9)
                c_v.fill = _fill(bg_p); c_v.alignment = _align("right")
            for col_n, val, base in [(6,p_ub,p_ingr),(9,p_uop,p_ingr),(11,p_un,p_ingr)]:
                pct_v = _pct(val, base)
                c_p = ws.cell(row, col_n, pct_v)
                c_p.number_format = FMT_PCT
                c_p.font = _font(bold=False, color="C0392B" if pct_v < 0 else "555555", size=9, italic=True)
                c_p.fill = _fill(bg_p); c_p.alignment = _align("right")
            row += 1

        # Fila TOTAL proyectos
        ws.row_dimensions[row].height = 15
        for col in range(1, NCOLS + 1):
            ws.cell(row, col).fill = _fill(C_TOT_ROW)
        c = ws.cell(row, 2, f"TOTAL  ({len(proyectos)} proyectos)")
        c.font = _font(bold=True, color="FFFFFF", size=9)
        c.fill = _fill(C_TOT_ROW); c.alignment = _align("left")
        for col_n, val in [(3,ingr),(4,cv_mat),(5,ut_bruta),(7,gastos_op),(8,ut_op),(10,ut_neta)]:
            c_v = ws.cell(row, col_n, val)
            c_v.number_format = FMT_MONEY
            c_v.font = _font(bold=True, color="FFD700" if val >= 0 else "FF9090", size=9)
            c_v.fill = _fill(C_TOT_ROW); c_v.alignment = _align("right")
        for col_n, val, base in [(6,ut_bruta,ingr),(9,ut_op,ingr),(11,ut_neta,ingr)]:
            pct_v = _pct(val, base)
            c_p = ws.cell(row, col_n, pct_v)
            c_p.number_format = FMT_PCT
            c_p.font = _font(bold=True, color="DDDDDD", size=9)
            c_p.fill = _fill(C_TOT_ROW); c_p.alignment = _align("right")
        row += 2  # spacer

    # ══════════════════════════════════════════════════════════════════════════
    # SECCIÓN 5 — RESUMEN POR CENTRO DE COSTO
    # ══════════════════════════════════════════════════════════════════════════
    cc_list = [c2 for c2 in value_cols_cc if c2 != "TOTAL"]
    if cc_list:
        ws.row_dimensions[row].height = 16
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
        c = ws.cell(row, 1, f"  RESUMEN POR CENTRO DE COSTO  ({len(cc_list)} centros)")
        c.font = _font(bold=True, color="FFFFFF", size=10)
        c.fill = _fill(C_SEC_HDR); c.alignment = _align("left")
        row += 1

        ws.row_dimensions[row].height = 14
        for col in range(1, NCOLS + 1):
            ws.cell(row, col).fill = _fill(C_TBL_HDR)
        for col_n, hdr, al in [
            (2,"Centro de Costo","left"),(3,"Ingresos Netos","right"),(4,"Costo Ventas","right"),
            (5,"Ut. Bruta","right"),(6,"% MB","right"),(7,"Gastos Op.","right"),
            (8,"Ut. Operacional","right"),(9,"% Op.","right"),(10,"Ut. Neta","right"),(11,"% Neta","right"),
        ]:
            c = ws.cell(row, col_n, hdr)
            c.font = _font(bold=True, color="FFFFFF", size=9)
            c.fill = _fill(C_TBL_HDR); c.alignment = _align(al)
        ws.cell(row, 1).fill = _fill(C_TBL_HDR)
        row += 1

        cc_data = []
        for cc in cc_list:
            cc_ingr = _sumcod(filas_cc, "4",       cc)
            cc_cmat = _sumcod(filas_cc, "5.1.1",   cc)
            cc_ub   = _sub(filas_cc,   "UTILIDAD BRUTA",       cc)
            cc_gop  = (_sumcod(filas_cc, "5.2.1.1", cc) +
                       _sumcod(filas_cc, "5.2.1.2", cc))
            cc_uop  = _sub(filas_cc,   "UTILIDAD OPERACIONAL", cc)
            cc_un   = _sub(filas_cc,   "UTILIDAD NETA",        cc)
            cc_data.append((cc, cc_ingr, cc_cmat, cc_ub, cc_gop, cc_uop, cc_un))
        cc_data.sort(key=lambda x: x[1], reverse=True)

        alt_c = False
        for cc, cc_ingr, cc_cmat, cc_ub, cc_gop, cc_uop, cc_un in cc_data:
            bg_c = C_ALT_LOC if alt_c else "FFFFFF"
            alt_c = not alt_c
            ws.row_dimensions[row].height = 13
            for col in range(1, NCOLS + 1):
                ws.cell(row, col).fill = _fill(bg_c)
            c = ws.cell(row, 2, cc)
            c.font = _font(bold=False, color="1A1A2E", size=9)
            c.fill = _fill(bg_c); c.alignment = _align("left")
            for col_n, val in [(3,cc_ingr),(4,cc_cmat),(5,cc_ub),(7,cc_gop),(8,cc_uop),(10,cc_un)]:
                c_v = ws.cell(row, col_n, val)
                c_v.number_format = FMT_MONEY
                c_v.font = _font(bold=False, color="C0392B" if val < 0 else "1A1A2E", size=9)
                c_v.fill = _fill(bg_c); c_v.alignment = _align("right")
            for col_n, val, base in [(6,cc_ub,cc_ingr),(9,cc_uop,cc_ingr),(11,cc_un,cc_ingr)]:
                pct_v = _pct(val, base)
                c_p = ws.cell(row, col_n, pct_v)
                c_p.number_format = FMT_PCT
                c_p.font = _font(bold=False, color="C0392B" if pct_v < 0 else "555555", size=9, italic=True)
                c_p.fill = _fill(bg_c); c_p.alignment = _align("right")
            row += 1

        # Totales CC
        cc_ingr_t = _sumcod(filas_cc, "4",       "TOTAL")
        cc_cmat_t = _sumcod(filas_cc, "5.1.1",   "TOTAL")
        cc_ub_t   = _sub(filas_cc,   "UTILIDAD BRUTA",       "TOTAL")
        cc_gop_t  = (_sumcod(filas_cc, "5.2.1.1", "TOTAL") +
                     _sumcod(filas_cc, "5.2.1.2", "TOTAL"))
        cc_uop_t  = _sub(filas_cc,   "UTILIDAD OPERACIONAL", "TOTAL")
        cc_un_t   = _sub(filas_cc,   "UTILIDAD NETA",        "TOTAL")

        ws.row_dimensions[row].height = 15
        for col in range(1, NCOLS + 1):
            ws.cell(row, col).fill = _fill(C_TOT_ROW)
        c = ws.cell(row, 2, f"TOTAL  ({len(cc_list)} centros de costo)")
        c.font = _font(bold=True, color="FFFFFF", size=9)
        c.fill = _fill(C_TOT_ROW); c.alignment = _align("left")
        for col_n, val in [(3,cc_ingr_t),(4,cc_cmat_t),(5,cc_ub_t),
                           (7,cc_gop_t),(8,cc_uop_t),(10,cc_un_t)]:
            c_v = ws.cell(row, col_n, val)
            c_v.number_format = FMT_MONEY
            c_v.font = _font(bold=True, color="FFD700" if val >= 0 else "FF9090", size=9)
            c_v.fill = _fill(C_TOT_ROW); c_v.alignment = _align("right")
        for col_n, val, base in [(6,cc_ub_t,cc_ingr_t),(9,cc_uop_t,cc_ingr_t),(11,cc_un_t,cc_ingr_t)]:
            pct_v = _pct(val, base)
            c_p = ws.cell(row, col_n, pct_v)
            c_p.number_format = FMT_PCT
            c_p.font = _font(bold=True, color="DDDDDD", size=9)
            c_p.fill = _fill(C_TOT_ROW); c_p.alignment = _align("right")

    ws.freeze_panes = "B4"


def _write_observaciones_sheet(ws, observaciones, empresa):
    """Escribe la pestaña de observaciones."""
    ws.row_dimensions[1].height = 22
    ws.merge_cells("A1:F1")
    c = ws.cell(1, 1, empresa)
    c.font = _font(bold=True, color=C_TITULO_FG, size=13)
    c.fill = _fill(C_TITULO_BG)
    c.alignment = _align("center")

    ws.merge_cells("A2:F2")
    c = ws.cell(2, 1, "Observaciones y Análisis de Estados Financieros")
    c.font = _font(bold=False, color=C_TITULO_FG, size=11)
    c.fill = _fill(C_HEADER_BG)
    c.alignment = _align("center")

    ws.row_dimensions[4].height = 20
    headers = ["#", "Categoría", "Nivel", "Observación"]
    col_widths = [5, 32, 12, 80]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(4, i, h)
        c.font = _font(bold=True, color=C_HEADER_FG, size=10)
        c.fill = _fill(C_HEADER_BG)
        c.alignment = _align("center" if i != 4 else "left")
        ws.column_dimensions[get_column_letter(i)].width = w

    nivel_colors = {"ALERTA": C_OBS_ALERTA, "AVISO": C_OBS_AVISO, "INFO": C_OBS_INFO}
    nivel_fg     = {"ALERTA": "922B21",      "AVISO": "7D6608",    "INFO": "1A5276"}

    for idx, obs in enumerate(observaciones, 1):
        row  = ws.max_row + 1
        ws.row_dimensions[row].height = 30
        nivel = obs.get("nivel", "INFO")
        bg    = nivel_colors.get(nivel, C_OBS_INFO)
        fg_n  = nivel_fg.get(nivel, "000000")

        ws.cell(row, 1, idx).font          = _font(size=9, color="555555")
        ws.cell(row, 1).fill               = _fill(bg)
        ws.cell(row, 1).alignment          = _align("center")

        ws.cell(row, 2, obs.get("categoria", "")).font = _font(bold=True, size=9, color=fg_n)
        ws.cell(row, 2).fill               = _fill(bg)
        ws.cell(row, 2).alignment          = _align("left", wrap=True)

        ws.cell(row, 3, nivel).font        = _font(bold=True, size=9, color=fg_n)
        ws.cell(row, 3).fill               = _fill(bg)
        ws.cell(row, 3).alignment          = _align("center")

        ws.cell(row, 4, obs.get("descripcion", "")).font = _font(size=9)
        ws.cell(row, 4).fill               = _fill(bg)
        ws.cell(row, 4).alignment          = _align("left", wrap=True)

    ws.freeze_panes = "A5"


def exportar_excel(
    empresa,
    filas_mes, value_cols_mes,
    filas_proyecto, value_cols_proyecto,
    filas_cc, value_cols_cc,
    observaciones,
    df_balance=None,
    df_mayor_completo=None,
    periodo_desc="",
    titulo_mes="Estado de Resultados Comparativo Mensual",
    titulo_proyecto="Estado de Resultados por Proyecto (MOD y CIF prorrateados por ingresos)",
    titulo_cc="Estado de Resultados por Centro de Costo",
    titulo_balance="Estado de Situación Financiera",
):
    wb = Workbook()

    # ── Pestaña 1: Dashboard ejecutivo ───────────────────────────────────────
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    _write_dashboard_sheet(
        ws_dash, empresa, periodo_desc,
        filas_mes, value_cols_mes,
        filas_proyecto, value_cols_proyecto,
        filas_cc, value_cols_cc,
        df_balance=df_balance,
    )

    # ── Pestaña 2: Por Mes — variación % vs período anterior ─────────────────
    ws_mes = wb.create_sheet("P&G por Mes")
    _write_pyg_sheet(
        ws_mes, empresa, titulo_mes,
        filas_mes, value_cols_mes,
        pct_mode="variation",
    )

    # ── Pestaña 3: Por Proyecto ──────────────────────────────────────────────
    ws_proy = wb.create_sheet("P&G por Proyecto")
    _write_pyg_sheet(ws_proy, empresa, titulo_proyecto, filas_proyecto, value_cols_proyecto)

    # ── Pestaña 4: Por Centro de Costo ───────────────────────────────────────
    ws_cc = wb.create_sheet("P&G por Centro de Costo")
    _write_pyg_sheet(ws_cc, empresa, titulo_cc, filas_cc, value_cols_cc)

    # ── Pestaña 5: Estado de Situación Financiera (Balance) ──────────────────
    if df_balance is not None and not df_balance.empty:
        ws_bg = wb.create_sheet("Situación Financiera")
        _write_balance_sheet(ws_bg, empresa, titulo_balance, df_balance)

    # ── Pestaña 6: Transacciones Sin Asignación ───────────────────────────────
    if df_mayor_completo is not None and not df_mayor_completo.empty:
        ws_sa = wb.create_sheet("Sin Asignación")
        _write_sin_asignacion_sheet(ws_sa, df_mayor_completo, empresa)

    # ── Pestaña 7: Observaciones ─────────────────────────────────────────────
    ws_obs = wb.create_sheet("Observaciones")
    _write_observaciones_sheet(ws_obs, observaciones, empresa)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
