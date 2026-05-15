"""
Genera el Excel de conciliación tributaria con 3 pestañas:
  1. Solo en SRI      → documentos del SRI no registrados en contabilidad
  2. Solo en Excel    → documentos contabilizados que no constan en el SRI
  3. Conciliados      → documentos presentes en ambos (con diferencia de valor)
"""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paleta ────────────────────────────────────────────────────────────────────
C_TITULO_BG  = "1A1A2E"
C_TITULO_FG  = "FFFFFF"
C_HEADER_BG  = "16213E"
C_HEADER_FG  = "FFFFFF"
C_ROW_A      = "FFFFFF"
C_ROW_B      = "EBF5FB"
C_SUBTOT_BG  = "154360"
C_SUBTOT_FG  = "F0F3FF"
C_ALERT_BG   = "FDECEA"
C_ALERT_FG   = "C0392B"
C_OK_BG      = "E8F5E9"
C_OK_FG      = "1B5E20"
C_WARN_BG    = "FFF8E1"
C_WARN_FG    = "E65100"
FMT_MONEY    = '#,##0.00'


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border():
    thin = Side(style="thin", color="DDDDDD")
    return Border(bottom=thin)

def _title_row(ws, text, n_cols):
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(1, 1, "BAEC BALDOSAS DEL ECUADOR S.A.S.")
    c.font = _font(bold=True, color=C_TITULO_FG, size=13)
    c.fill = _fill(C_TITULO_BG)
    c.alignment = _align("center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    c = ws.cell(2, 1, text)
    c.font = _font(bold=False, color=C_TITULO_FG, size=11)
    c.fill = _fill(C_HEADER_BG)
    c.alignment = _align("center")

def _header_row(ws, row_num, labels, widths):
    ws.row_dimensions[row_num].height = 22
    for i, (lbl, w) in enumerate(zip(labels, widths), 1):
        c = ws.cell(row_num, i, lbl)
        c.font      = _font(bold=True, color=C_HEADER_FG, size=10)
        c.fill      = _fill(C_HEADER_BG)
        c.alignment = _align("center")
        ws.column_dimensions[get_column_letter(i)].width = w

def _data_row(ws, row_num, values, alt, money_cols=(), color_cell: dict | None = None):
    ws.row_dimensions[row_num].height = 14
    bg = C_ROW_B if alt else C_ROW_A
    for col, val in enumerate(values, 1):
        override_bg = color_cell.get(col) if color_cell else None
        c = ws.cell(row_num, col, val)
        c.font      = _font(size=9, color=(color_cell.get(f"fg_{col}", "000000")) if color_cell else "000000")
        c.fill      = _fill(override_bg or bg)
        c.alignment = _align("right" if col in money_cols else "left")
        c.border    = _border()
        if col in money_cols and isinstance(val, (int, float)):
            c.number_format = FMT_MONEY

def _total_row(ws, row_num, n_cols, label, total, total_col):
    ws.row_dimensions[row_num].height = 16
    for col in range(1, n_cols + 1):
        val = label if col == 1 else (total if col == total_col else "")
        c = ws.cell(row_num, col, val)
        c.font      = _font(bold=True, color=C_SUBTOT_FG, size=10)
        c.fill      = _fill(C_SUBTOT_BG)
        c.alignment = _align("right" if col == total_col else "left")
        if col == total_col:
            c.number_format = FMT_MONEY


# ── Pestaña: Solo en SRI ─────────────────────────────────────────────────────

def _write_solo_sri(ws, df: pd.DataFrame, es_retencion: bool):
    cols_fac = ["RUC Emisor", "Razón Social", "Tipo", "Serie",
                "Fecha Emisión SRI", "Total SRI"]
    cols_ret = ["RUC Emisor (cliente)", "Razón Social", "Serie Ret.",
                "Clave Acceso", "Fecha Emisión SRI", "Nota"]
    cols = cols_ret if es_retencion else cols_fac
    widths_fac = [18, 38, 20, 22, 14, 14]
    widths_ret = [20, 38, 22, 52, 13, 30]
    widths = widths_ret if es_retencion else widths_fac

    n = len(cols)
    _title_row(ws, "Documentos en SRI  —  NO registrados en contabilidad", n)
    _header_row(ws, 3, cols, widths)
    ws.freeze_panes = "A4"

    money_cols = {6} if not es_retencion else set()
    alt = False
    for _, row in df.iterrows():
        r = ws.max_row + 1
        vals = [row.get(c, "") for c in cols]
        _data_row(ws, r, vals, alt, money_cols,
                  color_cell={1: C_ALERT_BG} if not es_retencion else None)
        alt = not alt

    if not es_retencion and not df.empty and "Total SRI" in df.columns:
        total = pd.to_numeric(df["Total SRI"], errors="coerce").sum()
        _total_row(ws, ws.max_row + 1, n, "TOTAL SIN REGISTRAR", total, 6)


# ── Pestaña: Solo en Excel ────────────────────────────────────────────────────

C_MANUAL_BG  = "EAF4EA"   # verde muy suave — facturas físicas (esperado no estar en SRI)
C_MANUAL_FG  = "1A5C1A"
C_ELECTR_BG  = "FFF3CD"   # amarillo — electrónica que debería estar en SRI
C_ELECTR_FG  = "7B4F00"


def _write_solo_excel(ws, df: pd.DataFrame, es_retencion: bool):
    cols_fac = ["Origen", "RUC Emisor", "Razón Social", "No. Comprobante",
                "Fecha Emisión", "Total Excel", "No. Autorización"]
    cols_ret = ["RUC Emisor (cliente)", "Razón Social", "No. Retención",
                "No. Factura", "Fecha"]
    cols = cols_ret if es_retencion else cols_fac
    widths_fac = [24, 18, 36, 26, 14, 14, 52]
    widths_ret = [20, 38, 26, 26, 14]
    widths = widths_ret if es_retencion else widths_fac

    n = len(cols)
    _title_row(ws, "Documentos en Excel  —  NO constan en el SRI", n)
    _header_row(ws, 3, cols, widths)
    ws.freeze_panes = "A4"

    money_cols = {6} if not es_retencion else set()
    alt = False
    for _, row in df.iterrows():
        r = ws.max_row + 1
        vals = [row.get(c, "") for c in cols]
        origen = str(row.get("Origen", ""))
        if "Manual" in origen or "Física" in origen:
            cc = {i: C_MANUAL_BG for i in range(1, n + 1)}
        elif "Electrónica" in origen:
            cc = {i: C_ELECTR_BG for i in range(1, n + 1)}
        else:
            cc = {i: C_WARN_BG for i in range(1, n + 1)} if not es_retencion else None
        _data_row(ws, r, vals, alt, money_cols, color_cell=cc)
        alt = not alt

    if not es_retencion and not df.empty and "Total Excel" in df.columns:
        total = pd.to_numeric(df["Total Excel"], errors="coerce").sum()
        _total_row(ws, ws.max_row + 1, n, "TOTAL NO EN SRI", total, 6)


# ── Pestaña: Conciliados ──────────────────────────────────────────────────────

def _write_conciliados(ws, df: pd.DataFrame, es_retencion: bool):
    cols_fac = ["RUC Emisor", "Razón Social", "Tipo", "Serie",
                "Fecha Emisión SRI",
                "Base SRI", "IVA SRI", "Total SRI",
                "Base Excel", "Total Excel",
                "Dif. Base", "Dif. Total", "Estado"]
    cols_ret = ["RUC Emisor (cliente)", "Razón Social",
                "Serie Ret. SRI", "Fecha Emisión SRI",
                "No. Ret. Excel", "Fac(s). Venta Excel", "Estado"]
    cols = cols_ret if es_retencion else cols_fac
    widths_fac = [18, 34, 18, 22, 14, 13, 10, 13, 13, 13, 12, 12, 20]
    widths_ret = [20, 34, 22, 13, 22, 24, 20]
    widths = widths_ret if es_retencion else widths_fac

    n = len(cols)
    _title_row(ws, "Documentos conciliados  —  Presentes en SRI y en Excel", n)
    _header_row(ws, 3, cols, widths)
    ws.freeze_panes = "A4"

    # Columnas con formato monetario (1-indexed)
    money_cols = {6, 7, 8, 9, 10, 11, 12} if not es_retencion else set()
    alt = False
    for _, row in df.iterrows():
        r = ws.max_row + 1
        vals = [row.get(c, "") for c in cols]
        estado = str(row.get("Estado", ""))
        if "fac. diferente" in estado.lower():
            bg_override = C_ALERT_BG   # rojo — retención aplicada a factura equivocada
        elif "Dif." in estado:
            bg_override = C_WARN_BG    # naranja — diferencia de valor
        elif "✅" in estado:
            bg_override = C_OK_BG      # verde — OK
        else:
            bg_override = None
        cc = {i: bg_override for i in range(1, n + 1)} if bg_override else None
        _data_row(ws, r, vals, alt, money_cols, color_cell=cc)
        alt = not alt

    if not es_retencion and not df.empty:
        def _sum(col): return pd.to_numeric(df.get(col, pd.Series()), errors="coerce").sum()
        t_base_sri = _sum("Base SRI");  t_total_sri = _sum("Total SRI")
        t_base_xl  = _sum("Base Excel"); t_total_xl  = _sum("Total Excel")
        # col positions: Base SRI=6, IVA SRI=7, Total SRI=8, Base XL=9, Total XL=10, Dif.B=11, Dif.T=12
        totals = {1: "TOTALES", 6: t_base_sri, 8: t_total_sri,
                  9: t_base_xl, 10: t_total_xl,
                  11: round(t_base_sri - t_base_xl, 2),
                  12: round(t_total_sri - t_total_xl, 2)}
        r = ws.max_row + 1
        ws.row_dimensions[r].height = 16
        for col in range(1, n + 1):
            val = totals.get(col, "")
            c = ws.cell(r, col, val)
            c.font      = _font(bold=True, color=C_SUBTOT_FG, size=10)
            c.fill      = _fill(C_SUBTOT_BG)
            c.alignment = _align("right" if col in money_cols else "left")
            if col in money_cols and isinstance(val, (int, float)):
                c.number_format = FMT_MONEY


# ── Exportador principal ──────────────────────────────────────────────────────

def exportar_conciliacion(resultados: dict, tipo_label: str) -> io.BytesIO:
    """
    resultados: dict con claves 'conciliados', 'solo_sri', 'solo_excel' (DataFrames).
    tipo_label: nombre del tipo de comprobante para las pestañas.
    Devuelve BytesIO con el Excel.
    """
    from modules.loader_tributario import TIPO_RETENCION
    es_ret = "retenci" in tipo_label.lower()

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Solo en SRI"
    _write_solo_sri(ws1, resultados.get("solo_sri", pd.DataFrame()), es_ret)

    ws2 = wb.create_sheet("Solo en Excel")
    _write_solo_excel(ws2, resultados.get("solo_excel", pd.DataFrame()), es_ret)

    ws3 = wb.create_sheet("Conciliados")
    _write_conciliados(ws3, resultados.get("conciliados", pd.DataFrame()), es_ret)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
