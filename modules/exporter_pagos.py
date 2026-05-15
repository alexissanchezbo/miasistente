"""
Genera el Excel de pago con 2 pestañas:
  1. Resumen de Pago  → Proveedor | Nro. Factura | Valor a Pagar
  2. Detalle Transferencias → ID | Valor | Tipo Cta | Nro. Cuenta | Glosa
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paleta (misma que exporter.py principal) ──────────────────────────────────
C_TITULO_BG  = "1A1A2E"
C_TITULO_FG  = "FFFFFF"
C_HEADER_BG  = "16213E"
C_HEADER_FG  = "FFFFFF"
C_ROW_A      = "FFFFFF"
C_ROW_B      = "EBF5FB"
C_SUBTOT_BG  = "154360"
C_SUBTOT_FG  = "F0F3FF"
FMT_MONEY    = '#,##0.00'


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border():
    thin = Side(style="thin", color="DDDDDD")
    return Border(bottom=thin)

def _header_row(ws, row_num, labels, widths):
    ws.row_dimensions[row_num].height = 22
    for i, (lbl, w) in enumerate(zip(labels, widths), 1):
        c = ws.cell(row_num, i, lbl)
        c.font    = _font(bold=True, color=C_HEADER_FG, size=10)
        c.fill    = _fill(C_HEADER_BG)
        c.alignment = _align("center")
        ws.column_dimensions[get_column_letter(i)].width = w

def _title_row(ws, text, n_cols):
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(1, 1, "BAEC BALDOSAS DEL ECUADOR S.A.S.")
    c.font = _font(bold=True, color=C_TITULO_FG, size=13)
    c.fill = _fill(C_TITULO_BG)
    c.alignment = _align("center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    c = ws.cell(2, 1, text)
    c.font = _font(bold=False, color=C_TITULO_FG, size=11)
    c.fill = _fill(C_HEADER_BG)
    c.alignment = _align("center")


# ── Pestaña 1: Resumen de Pago ────────────────────────────────────────────────

def _write_resumen(ws, filas):
    """
    filas: list of dicts con keys: proveedor, nro_factura, valor
    """
    n_cols = 3
    _title_row(ws, "Resumen de Pago a Proveedores", n_cols)

    labels = ["Nombre Proveedor", "Nro. Factura", "Valor a Pagar"]
    widths = [42, 28, 18]
    _header_row(ws, 3, labels, widths)
    ws.freeze_panes = "A4"

    proveedor_actual = None
    alt = False

    for fila in filas:
        prov = fila["proveedor"]
        row  = ws.max_row + 1
        ws.row_dimensions[row].height = 14

        # Subtotal de separación cuando cambia el proveedor
        if proveedor_actual is not None and prov != proveedor_actual:
            _write_subtotal_resumen(ws, proveedor_actual, filas)
            alt = False

        proveedor_actual = prov
        bg = C_ROW_B if alt else C_ROW_A
        alt = not alt

        def _cell(col, val, money=False):
            c = ws.cell(row, col, val)
            c.font = _font(size=9)
            c.fill = _fill(bg)
            c.alignment = _align("left" if col < 3 else "right")
            c.border = _border()
            if money:
                c.number_format = FMT_MONEY

        _cell(1, prov)
        _cell(2, fila["nro_factura"])
        _cell(3, fila["valor"], money=True)

    # Fila total general
    row = ws.max_row + 1
    ws.row_dimensions[row].height = 16
    total = sum(f["valor"] for f in filas)

    c1 = ws.cell(row, 1, "TOTAL A PAGAR")
    c1.font  = _font(bold=True, color=C_SUBTOT_FG, size=10)
    c1.fill  = _fill(C_SUBTOT_BG)
    c1.alignment = _align("left")

    c2 = ws.cell(row, 2, "")
    c2.fill = _fill(C_SUBTOT_BG)

    c3 = ws.cell(row, 3, total)
    c3.font   = _font(bold=True, color=C_SUBTOT_FG, size=10)
    c3.fill   = _fill(C_SUBTOT_BG)
    c3.alignment = _align("right")
    c3.number_format = FMT_MONEY


def _write_subtotal_resumen(ws, proveedor, filas):
    """No escribimos subtotales por proveedor en el resumen para mantenerlo limpio."""
    pass


# ── Pestaña 2: Detalle Transferencias ─────────────────────────────────────────

def _write_transferencias(ws, filas):
    """
    filas: list of dicts con keys:
        id_pago, valor, tipo_cta, nro_cuenta, glosa
    Solo incluye filas con datos bancarios completos.
    """
    n_cols = 5
    _title_row(ws, "Detalle de Transferencias Bancarias", n_cols)

    labels = ["ID (Cédula / RUC)", "Valor a Pagar", "Tipo Cta.", "Nro. Cuenta", "Glosa"]
    widths = [20, 16, 12, 24, 44]
    _header_row(ws, 3, labels, widths)
    ws.freeze_panes = "A4"

    alt = False
    for fila in filas:
        row = ws.max_row + 1
        ws.row_dimensions[row].height = 14
        bg = C_ROW_B if alt else C_ROW_A
        alt = not alt

        def _cell(col, val, money=False):
            c = ws.cell(row, col, val)
            c.font = _font(size=9)
            c.fill = _fill(bg)
            c.alignment = _align("left" if col != 2 else "right")
            c.border = _border()
            if money:
                c.number_format = FMT_MONEY

        _cell(1, fila["id_pago"])
        _cell(2, fila["valor"], money=True)
        _cell(3, fila["tipo_cta"])
        _cell(4, fila["nro_cuenta"])
        _cell(5, fila["glosa"])

    # Total
    row = ws.max_row + 1
    ws.row_dimensions[row].height = 16
    total = sum(f["valor"] for f in filas)

    for col in range(1, n_cols + 1):
        c = ws.cell(row, col, "TOTAL" if col == 1 else (total if col == 2 else ""))
        c.font  = _font(bold=True, color=C_SUBTOT_FG, size=10)
        c.fill  = _fill(C_SUBTOT_BG)
        if col == 2:
            c.number_format = FMT_MONEY
            c.alignment = _align("right")
        else:
            c.alignment = _align("left")


# ── Exportador principal ───────────────────────────────────────────────────────

def exportar_pagos(df_seleccionado):
    """
    Recibe el DataFrame de facturas seleccionadas (ya con columnas ID_PAGO, TIPO_CTA, NRO_CUENTA).
    Devuelve BytesIO con el Excel listo para descarga.
    """
    from modules.loader_pagos import _find_col

    col_prov   = _find_col(df_seleccionado.columns, "raz", "social") or "Razón Social"
    col_doc    = _find_col(df_seleccionado.columns, "documento") or "# Documento"
    # Preferir el valor editado por el usuario (_VALOR_PAGO) sobre el saldo original
    col_total  = "_VALOR_PAGO" if "_VALOR_PAGO" in df_seleccionado.columns else (
        _find_col(df_seleccionado.columns, "total") or "Total"
    )

    # Glosa: nombre proveedor + últimos 4 dígitos del nro de documento
    def build_glosa(row):
        prov = str(row.get(col_prov, "")).strip()
        doc  = str(row.get(col_doc,  "")).strip()
        sufijo = doc[-4:] if len(doc) >= 4 else doc
        return f"{prov} {sufijo}"

    # ── Datos Pestaña 1 ──────────────────────────────────────────────────────
    filas_resumen = []
    for _, row in df_seleccionado.iterrows():
        filas_resumen.append({
            "proveedor":   str(row.get(col_prov, "")).strip(),
            "nro_factura": str(row.get(col_doc,  "")).strip(),
            "valor":       float(row.get(col_total, 0) or 0),
        })
    filas_resumen.sort(key=lambda x: x["proveedor"])

    # ── Datos Pestaña 2 (solo con datos bancarios completos) ─────────────────
    filas_transf = []
    for _, row in df_seleccionado.iterrows():
        id_p  = str(row.get("ID_PAGO",    "")).strip()
        tipo  = str(row.get("TIPO_CTA",   "")).strip()
        nro   = str(row.get("NRO_CUENTA", "")).strip()

        # Solo incluir si tiene datos bancarios
        if nro and nro not in ("", "nan", "None"):
            filas_transf.append({
                "id_pago":    id_p,
                "valor":      float(row.get(col_total, 0) or 0),
                "tipo_cta":   tipo,
                "nro_cuenta": nro,
                "glosa":      build_glosa(row),
            })

    # ── Generar workbook ─────────────────────────────────────────────────────
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Resumen de Pago"
    _write_resumen(ws1, filas_resumen)

    ws2 = wb.create_sheet("Detalle Transferencias")
    _write_transferencias(ws2, filas_transf)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Exportador TXT Cash Management Banco Pichincha ────────────────────────────

def exportar_txt_pichincha(df_seleccionado) -> bytes:
    """
    Genera el archivo TXT tab-delimitado para carga en Cash Management Pichincha.
    Sin encabezado, 12 columnas por fila:
      ORIENTACIÓN | CONTRAPARTIDA | MONEDA | VALOR | FORMA DE PAGO |
      TIPO CTA | NUMERO CTA | REFERENCIA | TIPO ID | NUMERO ID |
      NOMBRE BENEFICIARIO | CODIGO BANCO

    Solo incluye filas con datos bancarios completos (NRO_CUENTA presente).
    REFERENCIA = primeras 2 palabras del proveedor + " FACT " + últimos 4 dígitos factura.
    VALOR = entero sin separadores; los 2 últimos dígitos son decimales (ej. $100.50 → "10050").
    """
    from modules.loader_pagos import _find_col
    from modules.bancos_ifis import buscar_codigo_banco

    col_prov  = _find_col(df_seleccionado.columns, "raz", "social") or "Razón Social"
    col_doc   = _find_col(df_seleccionado.columns, "#", "doc") or _find_col(df_seleccionado.columns, "documento") or "# Documento"
    # Preferir el valor editado por el usuario (_VALOR_PAGO) sobre el saldo original
    col_total = "_VALOR_PAGO" if "_VALOR_PAGO" in df_seleccionado.columns else (
        _find_col(df_seleccionado.columns, "total") or "Total"
    )
    col_banco = _find_col(df_seleccionado.columns, "banco") or "Banco"

    lines: list[str] = []

    for _, row in df_seleccionado.iterrows():
        nro_cta  = str(row.get("NRO_CUENTA", "")).strip()
        if not nro_cta or nro_cta in ("nan", "None", ""):
            continue  # Sin cuenta bancaria → no se puede incluir en la transferencia

        id_pago   = str(row.get("ID_PAGO",   "")).strip()
        tipo_cta  = str(row.get("TIPO_CTA",  "")).strip().upper()   # AHO / CTE
        valor     = float(row.get(col_total, 0) or 0)
        prov      = str(row.get(col_prov,    "")).strip()
        doc       = str(row.get(col_doc,     "")).strip()
        banco_nom = str(row.get(col_banco,   "")).strip()

        # ── ORIENTACIÓN (fijo) ────────────────────────────────────────────────
        orientacion = "PA"

        # ── CONTRAPARTIDA: ID del proveedor, max 20 chars ─────────────────────
        contrapartida = id_pago[:20] if id_pago else nro_cta[:20]

        # ── MONEDA (fija) ─────────────────────────────────────────────────────
        moneda = "USD"

        # ── VALOR: sin punto ni coma; últimos 2 dígitos = centavos ───────────
        valor_fmt = str(int(round(valor * 100)))

        # ── FORMA DE PAGO (fija = crédito a cuenta) ──────────────────────────
        forma_pago = "CTA"

        # ── TIPO CTA (AHO / CTE ya vienen correctos del loader) ──────────────
        tipo_cta_fmt = tipo_cta if tipo_cta in ("AHO", "CTE") else "AHO"

        # ── NUMERO CTA, max 20 ────────────────────────────────────────────────
        nro_cta_fmt = nro_cta[:20]

        # ── REFERENCIA ────────────────────────────────────────────────────
        # Pagos manuales traen _REFERENCIA ya lista; facturas la calculamos
        _ref_manual = str(row.get("_REFERENCIA", "")).strip()
        if _ref_manual and _ref_manual not in ("nan", "None"):
            referencia = _ref_manual[:40]
        else:
            palabras = prov.split()
            nombre_corto = " ".join(palabras[:2]) if len(palabras) >= 2 else prov
            sufijo_doc   = doc[-4:] if len(doc) >= 4 else doc
            referencia   = f"{nombre_corto} FACT {sufijo_doc}"[:40]

        # ── TIPO ID: C=cédula (10 dígitos), R=RUC (13 dígitos) ───────────────
        id_digits = "".join(c for c in id_pago if c.isdigit())
        tipo_id   = "C" if len(id_digits) == 10 else "R"

        # ── NUMERO ID, max 14 ─────────────────────────────────────────────────
        numero_id = id_pago[:14]

        # ── NOMBRE BENEFICIARIO, max 41 ───────────────────────────────────────
        nombre_ben = prov[:41]

        # ── CODIGO BANCO: lookup IFIS, 4 dígitos con ceros a la izquierda ────
        codigo_banco = buscar_codigo_banco(banco_nom)
        if codigo_banco:
            codigo_banco = codigo_banco.zfill(4)   # garantiza "0010" no "10"

        lines.append("\t".join([
            orientacion,
            contrapartida,
            moneda,
            valor_fmt,
            forma_pago,
            tipo_cta_fmt,
            nro_cta_fmt,
            referencia,
            tipo_id,
            numero_id,
            nombre_ben,
            codigo_banco,
        ]))

    return "\n".join(lines).encode("utf-8")
